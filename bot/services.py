from __future__ import annotations

import logging
from typing import List, Dict, Any
from pathlib import Path

from openpyxl import load_workbook

from .config import EXCEL_SERVICES_PATH
from . import yclients_api

logger = logging.getLogger(__name__)

DEFAULT_SERVICES = [
    {"specialist": "Марина", "category": "СПА", "name": "СПА Релакс"},
    {"specialist": "Марина", "category": "Массаж", "name": "Классический массаж"},
    {"specialist": "Елизавета", "category": "СПА Turkish", "name": "Turkish Classic"},
    {"specialist": "Ирина", "category": "Массаж", "name": "Лимфодренажный массаж"},
    {"specialist": "Татьяна", "category": "Арома программы", "name": "Арома Баланс"},
]

_SERVICES_CACHE: List[Dict[str, Any]] = []
_UNIQUE_SERVICES: List[Dict[str, Any]] = []
_UNIQUE_BY_ID: Dict[int, Dict[str, Any]] = {}


def _normalize_header(value: str) -> str:
    return value.strip().lower()


def _load_from_excel() -> List[Dict[str, Any]]:
    path = Path(EXCEL_SERVICES_PATH)
    if not path.exists():
        logger.warning("Excel services file not found: %s", path)
        return []

    wb = load_workbook(filename=path, read_only=True, data_only=True)
    if "services" in wb.sheetnames:
        ws = wb["services"]
    else:
        ws = wb.active

    rows = ws.iter_rows(values_only=True)
    try:
        headers = next(rows)
    except StopIteration:
        return []

    header_map = {}
    for idx, h in enumerate(headers):
        if not h:
            continue
        key = _normalize_header(str(h))
        header_map[key] = idx

    def col_index(*names: str) -> int | None:
        for name in names:
            idx = header_map.get(_normalize_header(name))
            if idx is not None:
                return idx
        return None

    specialist_col = col_index("specialist", "специалист")
    category_col = col_index("category", "категория", "type", "тип")
    service_col = col_index("service", "услуга")

    if specialist_col is None or category_col is None or service_col is None:
        logger.error("Excel header columns not found. Required: specialist/type/service")
        return []

    result = []
    for row in rows:
        specialist = row[specialist_col] if specialist_col < len(row) else None
        category = row[category_col] if category_col < len(row) else None
        service = row[service_col] if service_col < len(row) else None
        if not specialist or not category or not service:
            continue
        result.append(
            {
                "specialist": str(specialist).strip(),
                "category": str(category).strip(),
                "name": str(service).strip(),
            }
        )
    return result


def _ensure_cache() -> None:
    global _SERVICES_CACHE, _UNIQUE_SERVICES, _UNIQUE_BY_ID
    if _SERVICES_CACHE:
        return

    from_excel = _load_from_excel()
    raw = from_excel if from_excel else DEFAULT_SERVICES

    _SERVICES_CACHE = [
        {"id": idx + 1, **item}
        for idx, item in enumerate(raw)
    ]

    unique_map: Dict[tuple[str, str], Dict[str, Any]] = {}
    for service in _SERVICES_CACHE:
        key = (service["category"], service["name"])
        if key not in unique_map:
            unique_map[key] = {
                "category": service["category"],
                "name": service["name"],
                "specialists": set(),
            }
        unique_map[key]["specialists"].add(service["specialist"])

    _UNIQUE_SERVICES = []
    _UNIQUE_BY_ID = {}
    for idx, item in enumerate(sorted(unique_map.values(), key=lambda x: (x["category"], x["name"]))):
        entry = {
            "id": idx + 1,
            "category": item["category"],
            "name": item["name"],
            "specialists": sorted(item["specialists"]),
        }
        _UNIQUE_SERVICES.append(entry)
        _UNIQUE_BY_ID[entry["id"]] = entry


def get_categories(specialist_name: str | None = None) -> List[str]:
    _ensure_cache()
    services = _SERVICES_CACHE
    if specialist_name:
        services = [s for s in services if s["specialist"] == specialist_name]
    categories = sorted({s["category"] for s in services})
    return categories


def get_services_by_category(category: str, specialist_name: str | None = None) -> List[Dict[str, Any]]:
    _ensure_cache()
    services = _SERVICES_CACHE
    if specialist_name:
        services = [s for s in services if s["specialist"] == specialist_name]
    filtered = [s for s in services if s["category"] == category]
    if specialist_name:
        seen = set()
        unique = []
        for s in filtered:
            if s["name"] in seen:
                continue
            seen.add(s["name"])
            unique.append(s)
        return unique
    return filtered


def get_unique_services_by_category(category: str) -> List[Dict[str, Any]]:
    _ensure_cache()
    return [s for s in _UNIQUE_SERVICES if s["category"] == category]


def get_service_by_id(service_id: int) -> Dict[str, Any] | None:
    _ensure_cache()
    for service in _SERVICES_CACHE:
        if service["id"] == service_id:
            return service
    return None


def get_unique_service_by_id(service_id: int) -> Dict[str, Any] | None:
    _ensure_cache()
    return _UNIQUE_BY_ID.get(service_id)


def get_specialists_for_unique_service_id(service_id: int) -> List[str]:
    service = get_unique_service_by_id(service_id)
    if not service:
        return []
    return service["specialists"]


def get_specialists_for_service_id(service_id: int) -> List[str]:
    _ensure_cache()
    service = get_service_by_id(service_id)
    if not service:
        return []
    name = service["name"]
    return sorted({s["specialist"] for s in _SERVICES_CACHE if s["name"] == name})


def get_services_from_api() -> List[Dict[str, Any]]:
    try:
        return yclients_api.get_services()
    except Exception as exc:
        logger.warning("YCLIENTS get_services failed, using local data: %s", exc)
        _ensure_cache()
        return _SERVICES_CACHE
